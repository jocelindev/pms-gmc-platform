from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "kobo_forms" / "maj_20260803" / "PMS_GMC_Formulaire_2_Donnees_Calcul_Journalieres_2026_MAJ_20260803_KPI_COMPLETS_134_ID_AUTO.xlsx"
OUTPUT_DIR = ROOT / "kobo_forms" / "maj_20260810"
OUTPUT = OUTPUT_DIR / "PMS_GMC_Formulaire_3_Donnees_Calcul_Flexible_2026_V4.xlsx"


SURVEY_HEADERS = [
    "type",
    "name",
    "label",
    "hint",
    "required",
    "default",
    "calculation",
    "constraint",
    "constraint_message",
    "relevant",
    "appearance",
    "choice_filter",
]


def row(
    type_: str,
    name: str,
    label: str = "",
    hint: str = "",
    required: str = "",
    default: str = "",
    calculation: str = "",
    constraint: str = "",
    constraint_message: str = "",
    relevant: str = "",
    appearance: str = "",
    choice_filter: str = "",
) -> list[str]:
    return [
        type_,
        name,
        label,
        hint,
        required,
        default,
        calculation,
        constraint,
        constraint_message,
        relevant,
        appearance,
        choice_filter,
    ]


def rows_from_sheet(ws) -> list[list]:
    return [list(values) for values in ws.iter_rows(values_only=True) if any(value is not None for value in values)]


def append_choice(ws, list_name: str, name: str, label: str, pole_filter: str = "", kpi_filter: str = "") -> None:
    ws.append([list_name, name, label, pole_filter, kpi_filter])


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source = load_workbook(SOURCE)
    target = load_workbook(SOURCE)

    survey = target["survey"]
    survey.delete_rows(1, survey.max_row or 1)
    survey.append(SURVEY_HEADERS)

    survey_rows = [
        row("start", "start", "Debut"),
        row("end", "end", "Fin"),
        row("today", "date_de_soumission", "Date de soumission"),
        row(
            "select_one countries",
            "pays_filiale",
            "Pays / Filiale",
            "Choisir Groupe uniquement si la donnee est consolidee groupe.",
            "yes",
            "Groupe",
        ),
        row(
            "select_one poles",
            "pole_id",
            "Pole / Direction",
            "La liste des KPI est filtree selon le pole choisi.",
            "yes",
            "",
            "",
            "",
            "",
            "",
            "autocomplete",
        ),
        row(
            "select_one kpi_ids",
            "id_kpi",
            "ID KPI",
            "Choisir le KPI concerne par la donnee.",
            "yes",
            "",
            "",
            "",
            "",
            "",
            "autocomplete",
            "pole_filter=${pole_id}",
        ),
        row(
            "date",
            "date_collecte",
            "Date de la donnee",
            "Pour les donnees journalieres: saisir la date exacte. Pour hebdomadaire: saisir la date de fin de semaine.",
            "yes",
        ),
        row(
            "select_one modes_saisie_donnees",
            "mode_saisie_donnee",
            "Mode de saisie",
            "Choisir selon ce que vous possedez: le realise final ou les donnees de calcul.",
            "yes",
            "valeur_directe",
        ),
        row(
            "decimal",
            "valeur_realisee",
            "Taux realise / Vs Target direct",
            "A utiliser si vous connaissez directement le taux de realisation du KPI. La plateforme l'utilise comme Vs Target sans recalcul realise / objectif. Exemple: 91 pour 91%.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='valeur_directe'",
        ),
        row(
            "note",
            "note_elements",
            "Renseigner les donnees de calcul disponibles",
            "Remplir une, deux ou trois donnees selon la formule. Les lignes vides seront ignorees.",
            "",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul'",
        ),
        row(
            "select_one kpi_elements",
            "element_id_1",
            "Element de calcul 1",
            "Premier element utilise dans la formule.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul'",
            "autocomplete",
            "kpi_filter=${id_kpi}",
        ),
        row(
            "decimal",
            "valeur_element_1",
            "Valeur element 1",
            "Saisir uniquement une valeur numerique.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul'",
        ),
        row(
            "select_one yes_no",
            "ajouter_element_2",
            "Ajouter un 2e element ?",
            "",
            "yes",
            "non",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul'",
        ),
        row(
            "select_one kpi_elements",
            "element_id_2",
            "Element de calcul 2",
            "Deuxieme element utilise dans la formule.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul' and ${ajouter_element_2}='oui'",
            "autocomplete",
            "kpi_filter=${id_kpi}",
        ),
        row(
            "decimal",
            "valeur_element_2",
            "Valeur element 2",
            "Saisir uniquement une valeur numerique.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul' and ${ajouter_element_2}='oui'",
        ),
        row(
            "select_one yes_no",
            "ajouter_element_3",
            "Ajouter un 3e element ?",
            "",
            "yes",
            "non",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul' and ${ajouter_element_2}='oui'",
        ),
        row(
            "select_one kpi_elements",
            "element_id_3",
            "Element de calcul 3",
            "Troisieme element utilise dans la formule.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul' and ${ajouter_element_3}='oui'",
            "autocomplete",
            "kpi_filter=${id_kpi}",
        ),
        row(
            "decimal",
            "valeur_element_3",
            "Valeur element 3",
            "Saisir uniquement une valeur numerique.",
            "yes",
            "",
            "",
            "",
            "",
            "${mode_saisie_donnee}='elements_calcul' and ${ajouter_element_3}='oui'",
        ),
        row("text", "source_donnee", "Source de la donnee", "Exemple: rapport WFM, fichier production, extraction CRM.", ""),
        row("text", "responsable_repondant", "Responsable / repondant", "", ""),
        row("select_one validations", "validation_hierarchique", "Validation hierarchique", "", "yes", "a_valider"),
        row("text", "commentaires", "Commentaires", "", "", "", "", "", "", "", "multiline"),
        row("calculate", "periode_reporting", "", "", "", "", "${date_collecte}"),
        row("calculate", "cle_pms_donnee", "", "", "", "", "concat(${pays_filiale}, '|', ${pole_id}, '|', ${id_kpi}, '|', ${date_collecte})"),
    ]
    for item in survey_rows:
        survey.append(item)

    choices = target["choices"]
    existing = {
        (str(values[0]), str(values[1]))
        for values in choices.iter_rows(min_row=2, values_only=True)
        if values and values[0] is not None and values[1] is not None
    }
    for choice in [
        ("modes_saisie_donnees", "valeur_directe", "Je connais le taux realise / Vs Target du KPI"),
        ("modes_saisie_donnees", "elements_calcul", "Je renseigne les donnees de calcul de la formule"),
        ("yes_no", "oui", "Oui"),
        ("yes_no", "non", "Non"),
    ]:
        if (choice[0], choice[1]) not in existing:
            append_choice(choices, *choice)

    settings = target["settings"]
    settings_values = rows_from_sheet(settings)
    settings.delete_rows(1, settings.max_row or 1)
    if settings_values:
        for values in settings_values:
            settings.append(values)
    else:
        settings.append(["form_title", "form_id", "version", "default_language"])
    header = [cell.value for cell in settings[1]]
    if "form_title" in header:
        title_col = header.index("form_title") + 1
        settings.cell(row=2, column=title_col).value = "PMS GMC - Formulaire 3 - Donnees de calcul flexible"
    if "form_id" in header:
        id_col = header.index("form_id") + 1
        settings.cell(row=2, column=id_col).value = "pms_gmc_formulaire_3_donnees_calcul_flexible_v4"
    if "version" in header:
        version_col = header.index("version") + 1
        settings.cell(row=2, column=version_col).value = "20260810_v4"

    target.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
